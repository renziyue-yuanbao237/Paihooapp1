#1721822624232088780005005254858227864441470903953 //77777777**7
from openpyxl import Workbook
from openpyxl.styles import colors, fills, Font, PatternFill

#A* = name
#*1 = date
book = Workbook()
sheet = book.active
sequence = list(map(lambda x: chr(x), range(ord('A'), ord('Z') + 1)))


def ten2TwentySix(num):
    L = []
    if num > 25:
        while True:
            d = int(num / 26)
            remainder = num % 26
            if d <= 25:
                L.insert(0, sequence[remainder])
                L.insert(0, sequence[d - 1])
                break
            else:
                L.insert(0, sequence[remainder])
                num = d - 1
    else:
        L.append(sequence[num])
  
    return "".join(L)
def _roll(num):
    def f(n,x):
        li=[]
        a=[0,1,2,3,4,5,6,7,8,9,'A','b','C','D','E','F']
        b=[]
        while True:
            s=n//x
            y=n%x
            b=b+[y]
            if s==0:
                break
            n=s
        b.reverse()
        for i in b:
            li.append(str(a[i]))
        li="".join(li)
        return int(li)
    di=["星期日","星期一","星期二","星期三","星期四","星期五","星期六"]
    num=int(str(f(num,7))[::-1][0])
    return di[num]
def roll(startwith,n):
    for x in range(startwith,n+startwith):
        yield _roll(x)
def set_color(num):
    with open("nlist-config.txt","r") as f:
        c=1
        for x in f.read().split("\n"):
            print(x)
            if x != "normal":
                for y in range(num):
                    sheet[ten2TwentySix(y)+str(c)].fill=PatternFill(patternType=fills.FILL_SOLID,fgColor=x)
            c+=1
while True:
    if True:
        #Get Segment
        with open("nlist.txt","rb") as f:
            index=2
            for var1 in f.read().decode("utf-8").split("\n"):
                sheet["A"+str(index)]=var1
                index+=1
    if True:
        #Date Segment
        s=int(input("startwith:"))
        n=int(input("number:"))
        dt=roll(s,n)
        c=1
        for var2 in dt:
            sheet[ten2TwentySix(c)+"1"]=var2
            c+=1
        set_color(num=n)
    break

book.save("work.xlsx")
